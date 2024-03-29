import tempfile
import json
from zipfile import ZipFile
from openpyxl import load_workbook
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect, reverse
from django.conf import settings
from core import flash, flash_get_messages
from services import prefix_service, users_service, product_service
from .forms import PrefixActionForm, StartingNumberForm
from products.models import Product
from barcoding.utilities import normalize


def prefixes_list(request):
    current_user = request.user
    company_organisation = users_service.get_company_organisation(current_user)

    prefixes = prefix_service.all()
    susp_prefixes = prefix_service.find(company_organisation=company_organisation, is_suspended=True).all()

    '''
    prefixes = prefix_service.find_all().all()
    result = db.session.query('prefix', 'products'). \
        from_statement(text("select products.gs1_company_prefix as prefix, 
                                    count(*) as products
                             from products
                             where owner_id=%s
                             group by products.gs1_company_prefix" % current_user.id)).all()

    result_locations = db.session.query('prefix', 'locations'). \
        from_statement(text("select locations.gs1_company_prefix as prefix, 
                                    count(*) as locations
                             from locations
                             where owner_id=%s
                             group by locations.gs1_company_prefix" % current_user.id)).all()

    # set products count
    for prefix in prefixes:
        for row in result:
            if prefix.prefix == row[0]:
                setattr(prefix, 'products', row[1])

    # set locations count
    for prefix in prefixes:
        for row in result_locations:
            if prefix.prefix == row[0]:
                setattr(prefix, 'locations', row[1])
    '''

    if request.method == 'POST':
        form = PrefixActionForm(request.POST)
        form.fields['select_prefix'].choices = [(str(p.id), p.prefix) for p in prefixes]
        if form.is_valid():
            try:
                int_prefix_id = int(form.data['select_prefix'])
            except (ValueError, TypeError):
                flash(request, 'Your selections were not valid!', 'danger')
            else:
                prefix = prefix_service.get(int_prefix_id)
                if not prefix:
                    raise Http404('Prefix not found')

                if prefix.is_special == 'READ-ONLY' and form.data['prefix_action'] != 'set_this':
                    flash(request, 'Read-only prefix, please contact GS1 helpdesk.', 'danger')
                else:
                    prefix_service.make_active(prefix.id)

                    prefix_action = form.data['prefix_action']
                    # Enter a new product in selected range
                    if prefix_action == 'new_product':
                        return redirect(reverse('user:products.add_product') + '?prefix=' + str(prefix.prefix))

                    # Set selected range as active and go to My Products
                    elif prefix_action == 'set_this':
                        return redirect(reverse('user:products.products_list'))

                    # Set starting GTIN in selected range manually
                    elif prefix_action == 'starting_gtin':
                        return redirect(reverse('prefixes:prefixes_set_starting', args=(prefix.id,)))

                    # Set starting GTIN to first available number
                    elif prefix_action == 'first_available':
                        try:
                            prefix.make_starting_from()
                        except Exception as e:
                            return render(request, 'prefixes/prefix_exhausted.html',
                                                   {'current_user': current_user, 'prefix': prefix })
                        prefix_service.save(prefix)
                        flash(request, 'Starting gtin has been set to GTIN-%s' % prefix.starting_from, 'success')
                        return redirect(reverse('prefixes:prefixes_list'))

                    # new location
                    elif prefix_action == 'new_gln':
                        return redirect(reverse('user:locations.add_location') + '?prefix=' + str(prefix.prefix))

                    elif prefix_action == 'first_available_gln':
                        pass
                        '''
                        try:
                            prefix.make_starting_from_gln()
                        except Exception,e:
                            return render_template('prefixes/prefix_exhausted.html', prefix=prefix)
                        prefix_service.save(prefix)
                        return redirect(url_for('.prefixes_list'))
                        '''

                    # Export available GTINs in this range
                    elif prefix_action == 'export_available':
                        try:
                            products = ( Product.objects.filter(owner = current_user)
                                                        .filter(gs1_company_prefix = prefix.prefix)
                                                        .order_by('gtin') )
                            prfxs = prefix.get_available_gtins(products)
                            if len(prfxs) > 0:
                                tfile = tempfile.NamedTemporaryFile(suffix='.xlsx')
                                zfile = tempfile.NamedTemporaryFile(suffix='.zip')

                                file_xlsx = load_workbook(settings.PREFIXES_EXCEL_TEMPLATE)
                                ws = file_xlsx.active
                                for index, prfx in enumerate(prfxs):
                                    _ = ws.cell(column=2, row=index + 5, value=prfx)
                                file_xlsx.save(filename=tfile.name)

                                with ZipFile(zfile.name, "w") as z:
                                    export_filename = "export_%s_available.xlsx" % (prefix.prefix,)
                                    attachment_filename = "export_%s_available.%s" % (prefix.prefix, 'zip')
                                    z.write(tfile.name, export_filename)

                                send_file = open(zfile.name, 'rb')
                                response = HttpResponse(send_file, content_type='application/zip')
                                response['Content-Disposition'] = 'attachment; filename=%s' % attachment_filename
                                return response
                            else:
                                flash(request, 'There are no available GTIN numbers for current active prefix', 'danger')
                        except Exception as e:
                            flash(request, 'Error: %s' % str(e), 'danger')
        else:
            flash(request, 'You must choose a prefix and an action!', 'danger')

    form = PrefixActionForm()
    form.fields['select_prefix'].choices = [(str(p.id), p.prefix) for p in prefixes]
    try:
        selected_prefix = int(request.POST['select_prefix'])
        prefix_service.make_active(selected_prefix)
    except:
        selected_prefix = prefix_service.get_active(company_organisation)

    config = {'GS1_GLN_CAPABILITY': settings.GS1_GLN_CAPABILITY}

    context = {
        'current_user': current_user,
        'config': config,
        'prefixes': prefixes,
        'susp_prefixes': susp_prefixes,
        'flashed_messages': flash_get_messages(request),
        'selected_prefix': selected_prefix
    }
    return render(request, 'prefixes/prefixes_list.html', context)


def prefixes_set_starting(request, prefix_id):
    current_user = users_service.get(1)
    prefix = prefix_service.find_item(id=prefix_id)
    if not prefix:
        Http404()
    sn_length = 12 - len(prefix.prefix)
    if request.method == 'POST':
        error = 'Incorrect entry. Please enter a valid number'
        form = StartingNumberForm(request.POST)
        if form.is_valid():
            if len(form.data['starting_number']) == sn_length:
                try:
                    int(form.data['starting_number'])
                except (ValueError, TypeError):
                    pass
                else:
                    starting_number = normalize('EAN13', prefix.prefix + form.data['starting_number'])
                    products = product_service.find(gtin="0" + starting_number, owner=current_user).all()
                    if len(products) == 0:
                        prefix.starting_from = starting_number
                        prefix_service.save(prefix)
                        flash(request, 'Starting gtin has been set to GTIN-%s' % prefix.starting_from, 'success')
                        return redirect(reverse('prefixes:prefixes_list'))
                    else:
                        error = 'This number is already assigned. Try another one.'
        flash(request, error, 'danger')
    form = StartingNumberForm()
    if not prefix.starting_from:
        try:
            prefix.make_starting_from()
        except:
            return render(request, 'prefixes/prefix_exhausted.html',
                                   {'current_user': current_user, 'prefix': prefix})
        prefix_service.save(prefix)
    form.data['starting_number'] = prefix.starting_from[len(prefix.prefix):12]
    context = { 'current_user': current_user,
                'prefix': prefix,
                'form': form,
                'current': prefix.starting_from,
                'sn_length': sn_length,
                'flashed_messages': flash_get_messages(request) }
    return render(request, 'prefixes/set_starting.html', context=context)


def jsonify(**kwargs):
    content = json.dumps(kwargs)
    response = HttpResponse(content, content_type='application/json')
    response['Content-Length'] = len(content)
    return response


def prefixes_ajax(request):
    if request.method != 'POST':
        return Http404()
    current_user = request.user
    prefix_id = request.POST.get('pk', None)
    if not prefix_id:
        return Http404()
    company_organisation = users_service.get_company_organisation(current_user)
    prefix = prefix_service.find_item(company_organisation=company_organisation, id=prefix_id)
    if not prefix:
        return Http404()
    prefix.description = request.POST.get('value', None)
    prefix_service.save(prefix)
    return jsonify(success=True)
